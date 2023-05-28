import openpyxl, copy
from questionsparser import QuestionsParser
from typeform import Typeform


class Answer:

    def __init__(self, _ident=None, _levels=None, _text=None, _options=None):
        """levels is an array containing axis, domain, subdomain or more"""
        self.ident = _ident
        self.levels = _levels
        self.text = _text
        self.options = _options


class Candidate:
    def __init__(self, _name=None, _email=None):
        self.name = _name
        self.email = _email
        self.answers = []


class TypeFormResultsParser:

    @staticmethod
    def ident_from_text(text):
        return text[1:5]

    def __init__(self, _questions_parser=None, _count=74, _file="/Users/macluky/Downloads/responses.xlsx", _nr_candidates=1):
        self.file = _file
        self.count = _count
        self.questions_parser = _questions_parser
        self.candidates = []

        # generate this from typeform
        #if self.file is None:
        #self.api_key = "tfp_6da9A3giqGYanQRBWhY8oUkwFmokZiA5L1jTqVJvGxfP_iVx5bip5Ei3J"
        #typeform = Typeform(self.api_key)
        #forms: dict = typeform.forms.list()
        #print(forms)
        
        self.parse_excel_file(_nr_candidates)

    def parse_excel_file(self, _nr_candidates):
        workbook = openpyxl.load_workbook(self.file)
        sheet = workbook.active
        for row in range(2, 2 + _nr_candidates):
            # first col is a reference, 2 is name, 3 is email, then count questions
            name = sheet.cell(row=row, column=2).value
            email = sheet.cell(row=row, column=3).value
            candidate = Candidate(name, email)

            for column in range(4, 4 + self.count):
                q_text = sheet.cell(row=1, column=column).value
                q_answer = sheet.cell(row=row, column=column).value

                ident = self.ident_from_text(q_text)
                question = self.questions_parser.question_with_ident(ident)
                score = question.score_for_option(q_answer)
                answered_question = copy.deepcopy(question)
                answered_question.score = score

                candidate.answers.append(answered_question)
            self.candidates.append(candidate)


if __name__ == '__main__':
    qp = QuestionsParser()
    tp = TypeFormResultsParser(qp, 15, _nr_candidates=4)
    questions = tp.candidates
    print(questions)
