import openpyxl


class WeightParser:

    def __init__(self, _weight_tab="Baseline",
                 _file="/Users/macluky/Library/CloudStorage/OneDrive-SharedLibraries-ExpandiorAcademyB.V/Expandior Team - Documents/Product/Compass (Maturity Scan Personal)/Product Role Compass v4.9.xlsx"):
        self.file = _file
        self.weights = dict()

        workbook = openpyxl.load_workbook(self.file)
        sheet = workbook[_weight_tab]

        for row in range(2, 8):
            axis = sheet.cell(row=row, column=1).value
            area = sheet.cell(row=row, column=2).value
            weight = sheet.cell(row=row, column=3).value
            self.weights[(axis, area)] = weight

    def weight_label(self, label):
        # none is level 0
        if label is None:
            total = 0
            for w in self.weights.values():
                total += w
            return total/6
        else:
            count = 0
            total = 0
            for t in self.weights.keys():
                if label in t:
                    count += 1
                    total += self.weights[t]
            return total/count


if __name__ == '__main__':
    wp = WeightParser()
    print(wp.weights)
    print(wp.weight_label(None))
    print(wp.weight_label("Mastery"))

