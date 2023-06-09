import openpyxl


def is_valid_question(question, topic1, topic2, topic3, topic4, topic5, topic6):
    if topic1 == "" and "topic1" in question:
        return False
    if topic2 == "" and "topic2" in question:
        return False
    if topic3 == "" and "topic3" in question:
        return False
    if topic4 == "" and "topic4" in question:
        return False
    if topic5 == "" and "topic5" in question:
        return False
    if topic6 == "" and "topic6" in question:
        return False
    if question in ["topic1", "topic2", "topic3", "topic4", "topic5", "topic6", "professional", "company", "email",
                    "Start Date (UTC)", "Submit Date (UTC)", "Network ID"]:
        return False
    # print(question)
    return True


def replace_variables(key, professional, topic1, topic2, topic3, topic4, topic5, topic6):
    new_key = key.replace("{{hidden:professional}}", professional)
    new_key = new_key.replace("{{hidden:topic1}}", topic1)
    new_key = new_key.replace("{{hidden:topic2}}", topic2)
    new_key = new_key.replace("{{hidden:topic3}}", topic3)
    new_key = new_key.replace("{{hidden:topic4}}", topic4)
    new_key = new_key.replace("{{hidden:topic5}}", topic5)
    new_key = new_key.replace("{{hidden:topic6}}", topic6)
    return new_key


class FeedbackParser:

    def __init__(self, _file="/Users/macluky/Downloads/feedback_responses.xlsx"):
        self.file = _file
        self.professionals = dict()

        workbook = openpyxl.load_workbook(self.file, read_only=False)
        sheet = workbook.active

        for row in range(2, (sheet.max_row + 1)):
            responses = dict()
            email = None
            # extract the feedback from the feedback giver
            for column in range(3, sheet.max_column):
                key = sheet.cell(row=1, column=column).value
                value = sheet.cell(row=row, column=column).value
                responses[key] = value
                if key == "email":
                    email = value

            # find or create the professional
            if email in self.professionals.keys():
                professional = self.professionals[email]
            else:
                professional = []
                self.professionals[email] = professional

            # add the responses
            professional.append(responses)

    def have_feedback_for_email(self, email):
        return self.responses_for_email(email) is not None

    def responses_for_email(self, email):
        if email in self.professionals.keys():
            return self.professionals[email]
        else:
            #print("found no feedback for: " + email)
            return None

    def expanded_responses_for_email(self, email):
        responses = self.responses_for_email(email)
        expanded = []
        for response in responses:
            new_response = self.expand_response(response)
            expanded.append(new_response)

        return expanded

    def expand_response(self, response):

        professional = response["professional"]
        topic1 = response["topic1"]
        topic2 = response["topic2"]
        topic3 = response["topic3"]
        topic4 = response["topic4"]
        topic5 = response["topic5"]
        topic6 = response["topic6"]

        expanded = dict()

        for key in response.keys():
            if is_valid_question(key, topic1, topic2, topic3, topic4, topic5, topic6):
                value = response[key]
                new_key = replace_variables(key, professional, topic1, topic2, topic3, topic4, topic5, topic6)
                if key == "Are you willing to provide more details about your input after {{hidden:professional}} reviews your feedback?":
                    if value == '1':
                        value = "Yes"
                    else:
                        value = "No"
                expanded[new_key] = value
        return expanded


if __name__ == '__main__':
    fp = FeedbackParser()
    fp.responses_for_email("s@b.de")
    re = fp.responses_for_email("a.ganga@topdesk.com")
    er = fp.expanded_responses_for_email("a.ganga@topdesk.com")
