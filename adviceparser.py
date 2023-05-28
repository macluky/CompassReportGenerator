import openpyxl


class Advice:

    def __init__(self, _levels=None, _options=None):
        """levels is an array containing axis, domain, subdomain or more"""
        self.levels = _levels
        self.options = _options


class AdviceParser:

    def __init__(self, _count=18, _max_model_layers=3, _max_options=5,
                 _file="/Users/macluky/Library/CloudStorage/OneDrive-SharedLibraries-ExpandiorAcademyB.V/Expandior Team - Documents/Product/Compass (Maturity Scan Personal)/Product Role Compass v4.9.xlsx"):
        self.file = _file
        self.max_options = _max_options
        self.max_model_layers = _max_model_layers
        self.count = _count
        self.advices = []

        workbook = openpyxl.load_workbook(self.file)
        sheet = workbook["Advice"]

        for row in range(2, 2 + self.count):
            # extract the path of the levels (axis, area, subdomain)
            levels = []
            for model_layer in range(1, self.max_model_layers + 1):
                levels.append(sheet.cell(row=row, column=model_layer).value)
            options = []
            # parse each advice for the corresponding level
            for option in range(0, self.max_options):
                value = sheet.cell(row=row, column=4 + option).value
                if value is not None:
                    options.append(value)

            result = Advice(levels, options)
            self.advices.append(result)

    def advice_for_levels_and_reference(self, levels, ref):
        for result in self.advices:
            found_match = True
            for indent in range(0, len(levels)):
                if result.levels[indent] != levels[indent]:
                    found_match = False
            if found_match:
                return result.options[ref]

        print("Warning: no advice found for levels: " + levels)
        return None

    def advice_for_label_at_depth_with_reference(self, label, depth, ref):
        for result in self.advices:
            if result.levels[depth] == label:
                return result.options[int(ref)]

        print("Warning: no advice found for label: " + label)
        return None


if __name__ == '__main__':
    advices = AdviceParser().advices
    print(advices)
    advice = AdviceParser().advice_for_levels_and_reference(["Mastery", "Process", "Marketing"], 3)
    print(advice)
