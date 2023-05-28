import openpyxl


class Question:

    def __init__(self, _ident=None, _levels=None, _text=None, _options=None):
        """levels is an array containing axis, domain, subdomain or more"""
        self.ident = _ident
        self.levels = _levels
        self.text = _text
        self.options = _options
        self.score = None

    def score_for_option(self, answer):
        score = 0
        for option in self.options:
            if option.strip() == answer.strip():
                return score
            score += 1
        print("Warning: answer: [" + answer + "] not found in question: " + self.ident + " setting to 0")
        return 0

    def rank(self):
        return self.score / self.options.length


class QuestionsParser:

    def __init__(self, _count=74, _levels=3, _max_options=5,
                 _file="/Users/macluky/Library/CloudStorage/OneDrive-SharedLibraries-ExpandiorAcademyB.V/Expandior Team - Documents/Product/Compass (Maturity Scan Personal)/Product Role Compass v4.9.xlsx"):
        self.file = _file
        self.max_options = _max_options
        self.levels = _levels
        self.count = _count
        self.questions = []

        workbook = openpyxl.load_workbook(self.file)
        sheet = workbook["Questions"]

        for row in range(2, 2 + self.count):
            ident = sheet.cell(row=row, column=1).value
            levels = []
            for level in range(0, self.levels):
                levels.append(sheet.cell(row=row, column=2 + level).value)
            text = sheet.cell(row=row, column=5).value
            options = []
            for option in range(0, self.max_options):
                value = sheet.cell(row=row, column=6 + option).value
                if value is not None:
                    options.append(value)

            question = Question(ident, levels, text, options)
            self.questions.append(question)

    def question_with_ident(self, ident):
        for question in self.questions:
            if question.ident == ident:
                return question
        print("Warning: ident not found: "+ident)
        return None


if __name__ == '__main__':
    questions = QuestionsParser().questions
    print(questions)
